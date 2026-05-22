"""
将 人物对照表.xlsx + 提示词.xlsx 转换为 prompts.json

人物映射: [苏婉] → @suwan, [陈敬明] → @chenjingming 等
"""
import json
import re
import openpyxl
from pathlib import Path

BASE = Path(__file__).parent

# 1. 读取人物对照表
wb1 = openpyxl.load_workbook(BASE / "人物对照表.xlsx")
ws1 = wb1.active
char_map = {}
for row in ws1.iter_rows(min_row=2, values_only=True):  # 跳过表头
    prompt_name, ref_name = row[0], row[1]
    if prompt_name and ref_name:
        char_map[prompt_name.strip()] = ref_name.strip()

print("=== 人物映射 ===")
for k, v in char_map.items():
    print(f"  [{k}] → @{v}")

# 2. 读取提示词
wb2 = openpyxl.load_workbook(BASE / "提示词.xlsx")
ws2 = wb2.active

prompts = []
for row_idx, row in enumerate(ws2.iter_rows(min_row=2, values_only=True), start=2):
    raw_prompt = row[0] or ""
    duration_raw = str(row[1] or "5s").strip()
    ratio_raw = row[2]

    if not raw_prompt.strip():
        continue

    # 解析时长（支持 "12秒", "14s", "12S" 格式）
    dur_match = re.search(r'(\d+)', duration_raw)
    duration = int(dur_match.group(1)) if dur_match else 5
    # 限制在 Runway 支持的范围内
    duration = max(4, min(15, duration))

    # 解析比例 (datetime.time 对象 → "9:16")
    if hasattr(ratio_raw, 'strftime'):
        ratio = f"{ratio_raw.hour}:{ratio_raw.minute:02d}"  # 如 "9:16"
    else:
        ratio = str(ratio_raw or "16:9").strip()
        if ':' not in ratio:
            ratio = "16:9"

    # 提取所有 prompt: 后面的内容（核心提示词）
    prompt_lines = re.findall(r'prompt:\s*(.+?)(?:\n\n|$)', raw_prompt, re.DOTALL)
    if prompt_lines:
        # 合并多个 prompt 为一段
        core_prompt = " ".join(p.strip() for p in prompt_lines)
    else:
        # 没有 prompt: 标记，用全文
        core_prompt = raw_prompt

    # 替换 [人物名] → @引用名 (前后加空格以便 RunwayML 识别)
    for char_name, ref_name in char_map.items():
        core_prompt = core_prompt.replace(f"[{char_name}]", f" @{ref_name} ")

    # 识别引用的 references 列表
    references = []
    found_refs = set()
    for char_name, ref_name in char_map.items():
        if f"@{ref_name}" in core_prompt:
            if ref_name not in found_refs:
                references.append(ref_name)
                found_refs.add(ref_name)

    prompts.append({
        "references": references if references else [],
        "prompt": core_prompt.strip(),
        "duration": duration,
        "ratio": ratio,
        "_source_row": row_idx,
    })

print(f"\n=== 转换结果: {len(prompts)} 条 ===")
for i, p in enumerate(prompts):
    print(f"\n[{i+1}] refs={p['references']}  dur={p['duration']}s  ratio={p['ratio']}")
    print(f"    prompt: {p['prompt'][:120]}...")

# 3. 保存
output = BASE / "prompts.json"
output.write_text(json.dumps(prompts, indent=2, ensure_ascii=False), encoding="utf-8")
print(f"\n已保存 {len(prompts)} 条到: {output}")
