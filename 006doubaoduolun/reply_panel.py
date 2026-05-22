import os
import re

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem,
    QHeaderView, QAbstractItemView, QFileDialog, QApplication
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QFont
from qfluentwidgets import (
    PushButton, PrimaryPushButton, ToolButton, SubtitleLabel,
    FluentIcon, isDarkTheme
)

from models import ReplyMessage, ChatMode, SendMessage
from dialogs import DetailDialog

COL_ID = 0
COL_SEND_ID = 1
COL_CONTENT = 2
COL_TIME = 3
COL_ELAPSED = 4
COL_MODE = 5
COL_OPS = 6


class ReplyPanel(QWidget):
    reply_selected = Signal(int)  # emits send_id for cross-highlight

    def __init__(self, parent=None):
        super().__init__(parent)
        self._replies: list[ReplyMessage] = []
        self._get_send_messages = None  # callable -> list[SendMessage]
        self._build_ui()

    def set_send_source(self, get_messages):
        """Provide a callable that returns the current list of SendMessage."""
        self._get_send_messages = get_messages

    def _build_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(8)

        layout.addWidget(SubtitleLabel("豆包回复记录"))

        self.table = QTableWidget(0, 7)
        self.table.setHorizontalHeaderLabels(["序号", "关联", "回复内容", "采集时间", "耗时", "模式", "操作"])
        self.table.horizontalHeader().setSectionResizeMode(COL_CONTENT, QHeaderView.ResizeMode.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(COL_OPS, QHeaderView.ResizeMode.Fixed)
        self.table.setColumnWidth(COL_ID, 50)
        self.table.setColumnWidth(COL_SEND_ID, 50)
        self.table.setColumnWidth(COL_TIME, 90)
        self.table.setColumnWidth(COL_ELAPSED, 60)
        self.table.setColumnWidth(COL_MODE, 90)
        self.table.setColumnWidth(COL_OPS, 100)
        self.table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)
        self.table.setAlternatingRowColors(True)
        self.table.doubleClicked.connect(self._on_double_click)
        self.table.selectionModel().selectionChanged.connect(self._on_selection)
        layout.addWidget(self.table)

        btn_row = QHBoxLayout()
        export_btn = PushButton(FluentIcon.SAVE, "导出回复")
        export_btn.clicked.connect(self._export)
        clear_btn = PushButton(FluentIcon.DELETE, "清空记录")
        clear_btn.clicked.connect(self._clear)
        btn_row.addWidget(export_btn)
        btn_row.addWidget(clear_btn)
        btn_row.addStretch()
        layout.addLayout(btn_row)

    # ------------------------------------------------------------------ #
    #  Public API                                                          #
    # ------------------------------------------------------------------ #

    def add_reply(self, reply: ReplyMessage):
        self._replies.append(reply)
        self._append_row(reply)
        self.table.scrollToBottom()

    def highlight_by_send_id(self, send_id: int):
        for row, reply in enumerate(self._replies):
            if reply.send_id == send_id:
                self.table.selectRow(row)
                self.table.scrollTo(self.table.model().index(row, 0))
                break

    def clear(self):
        self._replies.clear()
        self.table.setRowCount(0)

    # ------------------------------------------------------------------ #
    #  Slots                                                               #
    # ------------------------------------------------------------------ #

    def _on_double_click(self, index):
        row = index.row()
        if row < len(self._replies):
            r = self._replies[row]
            DetailDialog(f"回复 #{r.id} 详情", r.content, self).exec()

    def _on_selection(self):
        rows = self.table.selectedItems()
        if rows:
            row = self.table.currentRow()
            if row < len(self._replies):
                self.reply_selected.emit(self._replies[row].send_id)

    def _copy_reply(self, row: int):
        if row < len(self._replies):
            QApplication.clipboard().setText(self._replies[row].content)

    def _view_detail(self, row: int):
        if row < len(self._replies):
            r = self._replies[row]
            DetailDialog(f"回复 #{r.id} 详情", r.content, self).exec()

    def _clear(self):
        from qfluentwidgets import MessageBox
        box = MessageBox("确认清空", "确定要清空全部回复记录吗？", self)
        if box.exec():
            self.clear()

    def _export(self):
        if not self._replies:
            return

        from PySide6.QtWidgets import QDialog, QVBoxLayout, QHBoxLayout
        from qfluentwidgets import PushButton, SubtitleLabel

        dlg = QDialog(self)
        dlg.setWindowTitle("选择导出格式")
        dlg.setFixedSize(300, 140)
        layout = QVBoxLayout(dlg)
        layout.setContentsMargins(24, 20, 24, 20)
        layout.setSpacing(16)
        layout.addWidget(SubtitleLabel("请选择导出格式"))
        btn_row = QHBoxLayout()
        btn_row.setSpacing(12)
        txt_btn = PushButton("TXT 文本")
        xlsx_btn = PushButton("Excel 表格")
        txt_btn.setFixedHeight(36)
        xlsx_btn.setFixedHeight(36)
        btn_row.addWidget(txt_btn)
        btn_row.addWidget(xlsx_btn)
        layout.addLayout(btn_row)

        chosen = [None]
        txt_btn.clicked.connect(lambda: (chosen.__setitem__(0, "txt"), dlg.accept()))
        xlsx_btn.clicked.connect(lambda: (chosen.__setitem__(0, "xlsx"), dlg.accept()))
        dlg.exec()

        if not chosen[0]:
            return

        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        if chosen[0] == "txt":
            path, _ = QFileDialog.getSaveFileName(
                self, "导出为 TXT", os.path.join(desktop, "doubao_replies.txt"),
                "文本文件 (*.txt)"
            )
            if path:
                self._export_txt(path)
        else:
            path, _ = QFileDialog.getSaveFileName(
                self, "导出为 Excel", os.path.join(desktop, "doubao_replies.xlsx"),
                "Excel 表格 (*.xlsx)"
            )
            if path:
                self._export_xlsx(path)

    def _export_txt(self, path: str):
        with open(path, "w", encoding="utf-8") as f:
            for r in self._replies:
                content = re.sub(r"\s+", " ", r.content).strip()
                f.write(f"{r.id}. {content}\n")

    def _export_xlsx(self, path: str):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Build lookup: send_id -> SendMessage
        send_map = {}
        if self._get_send_messages:
            send_map = {m.id: m for m in self._get_send_messages()}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "豆包回复记录"

        headers = ["发送序号", "发送消息内容", "发送模式", "回复序号", "回复内容"]
        header_fill = PatternFill("solid", fgColor="4FC3F7")
        header_font = Font(bold=True, color="FFFFFF")
        thin = Side(style="thin", color="CCCCCC")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for row_idx, r in enumerate(self._replies, 2):
            send_msg = send_map.get(r.send_id)
            send_content = send_msg.content if send_msg else ""
            send_mode = send_msg.mode.value if send_msg else ""
            if send_msg and send_msg.forced_mode and send_msg.forced_mode != ChatMode.AUTO:
                send_mode = send_msg.forced_mode.value

            row_data = [
                r.send_id,
                send_content,
                send_mode,
                r.id,
                r.content,
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=(col in (2, 5)))
                if col in (1, 3, 4):
                    cell.alignment = Alignment(horizontal="center", vertical="top")

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 60
        ws.row_dimensions[1].height = 24

        wb.save(path)

    # ------------------------------------------------------------------ #
    #  Table helpers                                                       #
    # ------------------------------------------------------------------ #

    def _append_row(self, reply: ReplyMessage):
        row = self.table.rowCount()
        self.table.insertRow(row)
        self.table.setItem(row, COL_ID, self._cell(str(reply.id), center=True))
        self.table.setItem(row, COL_SEND_ID, self._cell(f"#{reply.send_id}", center=True))
        preview = reply.content[:50] + ("..." if len(reply.content) > 50 else "")
        self.table.setItem(row, COL_CONTENT, self._cell(preview))
        self.table.setItem(row, COL_TIME, self._cell(reply.collect_time.strftime("%H:%M:%S"), center=True))
        self.table.setItem(row, COL_ELAPSED, self._cell(f"{reply.elapsed_seconds}秒", center=True))
        self.table.setItem(row, COL_MODE, self._cell(reply.mode.value, center=True))

        ops_widget = QWidget()
        ops_layout = QHBoxLayout(ops_widget)
        ops_layout.setContentsMargins(2, 2, 2, 2)
        ops_layout.setSpacing(2)

        from PySide6.QtCore import QSize
        copy_btn = ToolButton(FluentIcon.COPY)
        copy_btn.setFixedSize(QSize(28, 28))
        copy_btn.setToolTip("复制全文")
        copy_btn.clicked.connect(lambda _, r=row: self._copy_reply(r))

        view_btn = ToolButton(FluentIcon.VIEW)
        view_btn.setFixedSize(QSize(28, 28))
        view_btn.setToolTip("查看详情")
        view_btn.clicked.connect(lambda _, r=row: self._view_detail(r))

        for b in (copy_btn, view_btn):
            ops_layout.addWidget(b)
        self.table.setCellWidget(row, COL_OPS, ops_widget)
        self.table.setRowHeight(row, 36)

    @staticmethod
    def _cell(text: str, center: bool = False) -> QTableWidgetItem:
        item = QTableWidgetItem(text)
        item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        if center:
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
        return item
