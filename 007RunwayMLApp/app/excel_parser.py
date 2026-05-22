import logging
import re
import shutil
import tempfile
from pathlib import Path
import openpyxl

from .models import PromptItem
from .config import MIN_DURATION, MAX_DURATION

logger = logging.getLogger(__name__)


def _read_workbook_safe(excel_path: str) -> openpyxl.Workbook:
    """Load workbook safely — copies to temp if file is locked (e.g. open in Excel)."""
    try:
        return openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except (PermissionError, OSError):
        logger.info("文件被占用，复制到临时目录读取: %s", excel_path)
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        try:
            shutil.copy2(excel_path, tmp.name)
            return openpyxl.load_workbook(tmp.name, read_only=True, data_only=True)
        finally:
            try:
                Path(tmp.name).unlink()
            except OSError:
                pass


class ExcelParser:
    """Port of convert_excel.py logic. Parses 人物对照表.xlsx and 提示词.xlsx."""

    @staticmethod
    def parse_character_mapping(excel_path: str) -> dict[str, str]:
        """Read 人物对照表.xlsx, return {prompt_name: ref_name}."""
        try:
            wb = _read_workbook_safe(excel_path)
            ws = wb.active
            char_map = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                prompt_name, ref_name = row[0], row[1]
                if prompt_name and ref_name:
                    char_map[prompt_name.strip()] = ref_name.strip()
            wb.close()
            return char_map
        except Exception:
            logger.exception("解析人物对照表失败: %s", excel_path)
            raise

    @staticmethod
    def parse_prompts(
        excel_path: str,
        char_map: dict[str, str],
        prefix: str = "",
        suffix: str = "",
    ) -> list[PromptItem]:
        """Read 提示词.xlsx, substitute characters, return PromptItem list."""
        try:
            wb = _read_workbook_safe(excel_path)
            ws = wb.active
            items = []

            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=0):
                try:
                    raw_cell = str(row[0] or "")
                    duration_raw = str(row[1] or "5s").strip()
                    ratio_raw = row[2]

                    if not raw_cell.strip():
                        continue

                    dur_match = re.search(r"(\d+)", duration_raw)
                    duration = int(dur_match.group(1)) if dur_match else 5
                    duration = max(MIN_DURATION, min(MAX_DURATION, duration))

                    if hasattr(ratio_raw, "strftime"):
                        ratio = f"{ratio_raw.hour}:{ratio_raw.minute:02d}"
                    else:
                        ratio = str(ratio_raw or "16:9").strip()
                        if ":" not in ratio:
                            ratio = "16:9"

                    # Use the full cell content as the API prompt (not just prompt: lines)
                    api_prompt = raw_cell.strip()

                    # Detect references from [角色名] bracket patterns in raw text
                    references = []
                    found = set()
                    bracket_names = re.findall(r"\[([^\]]+)\]", raw_cell)
                    for name in bracket_names:
                        if name in char_map:
                            ref = char_map[name]
                            if ref not in found:
                                references.append(ref)
                                found.add(ref)

                    # Also scan for char_map keys appearing as plain text (no brackets needed)
                    for cn_name, ref_name in char_map.items():
                        if ref_name not in found and cn_name in raw_cell:
                            references.append(ref_name)
                            found.add(ref_name)

                    # Append reference suffix at end of prompt: "角色名是 @refname"
                    if references:
                        ref_parts = []
                        for ref in references:
                            char_name = next((cn for cn, rn in char_map.items() if rn == ref), None)
                            if char_name:
                                ref_parts.append(f"{char_name}是 @{ref}")
                        if ref_parts:
                            api_prompt = api_prompt + "\n\n" + " ,".join(ref_parts)

                    items.append(PromptItem(
                        index=idx,
                        prompt_text=api_prompt,
                        raw_prompt=raw_cell,
                        references=references,
                        duration=duration,
                        ratio=ratio,
                        source_row=idx + 2,
                        prefix=prefix,
                        suffix=suffix,
                    ))
                except Exception:
                    logger.exception("解析提示词第 %d 行失败", idx + 2)

            wb.close()
            return items
        except Exception:
            logger.exception("解析提示词文件失败: %s", excel_path)
            raise

    @staticmethod
    def write_status_back(excel_path: str, items: list[PromptItem]):
        """Write status and error back to 提示词.xlsx columns D and E."""
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active

            if ws.cell(row=1, column=4).value is None:
                ws.cell(row=1, column=4, value="状态")
            if ws.cell(row=1, column=5).value is None:
                ws.cell(row=1, column=5, value="失败原因")

            for item in items:
                row = item.source_row
                if row < 2:
                    continue
                ws.cell(row=row, column=4, value=item.status.value)
                ws.cell(row=row, column=5, value=item.error_message or "")

            wb.save(excel_path)
            wb.close()
        except (PermissionError, OSError):
            logger.warning("写回状态失败 — Excel 文件被占用，请关闭 Excel 后重试: %s", excel_path)
        except Exception:
            logger.exception("写回状态到 Excel 失败: %s", excel_path)

    @staticmethod
    def load_character_assets(json_path: str) -> dict:
        """Load character_assets.json, return {ref_name: CharacterAsset}."""
        try:
            import json
            from .models import CharacterAsset

            data = json.loads(Path(json_path).read_text(encoding="utf-8"))
            return {
                name: CharacterAsset(ref_name=name, asset_id=v["assetId"], url=v["url"])
                for name, v in data.items()
            }
        except Exception:
            logger.exception("加载角色资源文件失败: %s", json_path)
            raise
