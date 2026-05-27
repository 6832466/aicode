"""Excel 导出 —— 移植自 utils/excel_export.py"""
from __future__ import annotations

import hashlib

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

COLUMNS = ["序号", "标题", "类型", "时长", "发布状态", "时间", "浏览数", "评论数", "点赞数"]
COL_WIDTHS = [8, 40, 18, 10, 12, 22, 12, 12, 12]


def export_to_excel(all_data: dict, output_path: str) -> int:
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    total = 0
    for cat_name, cards in all_data.items():
        if not cards:
            continue

        base = cat_name.replace("高光-", "").replace("/", "-")
        sheet_name = base[:31] if len(base) <= 31 else base[:27] + "_" + hashlib.md5(cat_name.encode()).hexdigest()[:4]
        ws = wb.create_sheet(title=sheet_name)

        for ci, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        for ri, card in enumerate(cards, 2):
            for ci, col_name in enumerate(COLUMNS, 1):
                val = card.get(col_name, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = cell_align
                cell.border = thin_border
                cell.font = Font(name="微软雅黑", size=10)
            total += 1

        for ci, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    if len([v for v in all_data.values() if v]) > 1:
        summary_ws = wb.create_sheet(title="汇总", index=0)
        for ci, col_name in enumerate(COLUMNS, 1):
            cell = summary_ws.cell(row=1, column=ci, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        row_idx = 2
        for cat_name, cards in all_data.items():
            for card in cards:
                for ci, col_name in enumerate(COLUMNS, 1):
                    val = card.get(col_name, "")
                    cell = summary_ws.cell(row=row_idx, column=ci, value=val)
                    cell.alignment = cell_align
                    cell.border = thin_border
                    cell.font = Font(name="微软雅黑", size=10)
                row_idx += 1

        for ci, w in enumerate(COL_WIDTHS, 1):
            summary_ws.column_dimensions[summary_ws.cell(row=1, column=ci).column_letter].width = w

    wb.save(output_path)
    return total
