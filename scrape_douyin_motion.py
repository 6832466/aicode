"""抓取短剧版权平台 - 漫剧列表，输出到桌面 WPS Excel"""
import os
import sys
import time
from pathlib import Path

# 修复 Windows 控制台编码问题，启用无缓冲输出
try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

TARGET_URL = "https://www.shortdramas.com/page/copyright/book-manage?tab=motion_comic"
USER_DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "playwright_profile")
DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
OUTPUT_FILE = os.path.join(DESKTOP, "漫剧列表.xlsx")

LOGIN_TIMEOUT = 600  # 10分钟


def is_on_login_page(page):
    """检测是否在登录页面"""
    url = page.url.lower()
    return any(kw in url for kw in ["login", "signin", "auth", "passport"])


def wait_for_table(page, timeout=15):
    """等待表格出现，返回是否成功"""
    try:
        page.wait_for_selector("table tbody tr", timeout=timeout * 1000)
        return True
    except Exception:
        return False


def scrape_page(page):
    """抓取当前页面上的漫剧列表数据"""
    rows = []
    page.wait_for_timeout(2000)

    try:
        page.wait_for_selector("table tbody tr", timeout=15000)
        trs = page.query_selector_all("table tbody tr")
        print(f"  找到 {len(trs)} 行表格数据")

        for tr in trs:
            tds = tr.query_selector_all("td")
            row_data = [td.inner_text().strip() for td in tds]
            if row_data:
                rows.append(row_data)
    except Exception as e:
        print(f"  表格抓取失败: {e}")

    if not rows:
        print("  尝试从页面文本提取...")
        try:
            body = page.inner_text("body")
            print(f"  页面文本长度: {len(body)} 字符")
            debug_file = os.path.join(DESKTOP, "page_debug.txt")
            with open(debug_file, "w", encoding="utf-8") as f:
                f.write(body)
            print(f"  页面文本已保存到: {debug_file}")
        except Exception as e:
            print(f"  文本提取失败: {e}")

    return rows


def save_to_excel(all_rows, filepath):
    """保存数据到 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "漫剧列表"

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    if all_rows:
        headers = all_rows[0]
        data = all_rows[1:]
    else:
        headers = ["序号", "内容"]
        data = []

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_align
            cell.border = thin_border
            cell.font = Font(name="微软雅黑", size=10)

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                max_len = max(max_len, length)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    wb.save(filepath)
    print(f"\n[数据] 已保存到: {filepath}")
    print(f"   共 {len(data)} 条记录")


def main():
    print("=" * 60, flush=True)
    print("  短剧版权平台 - 漫剧列表采集工具", flush=True)
    print("=" * 60, flush=True)

    os.makedirs(USER_DATA_DIR, exist_ok=True)

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=USER_DATA_DIR,
            headless=False,
            args=["--start-maximized"],
            no_viewport=True,
        )

        page = context.new_page()
        page.set_default_timeout(15000)

        print(f"\n正在打开页面: {TARGET_URL}", flush=True)

        try:
            page.goto(TARGET_URL, wait_until="domcontentloaded", timeout=60000)
        except Exception as e:
            print(f"页面加载警告: {e}", flush=True)

        # 等待页面渲染
        page.wait_for_timeout(5000)
        print(f"当前URL: {page.url}", flush=True)

        # 检测是否需要登录
        if is_on_login_page(page):
            print("\n[!] 检测到登录页面，请在浏览器中完成登录...", flush=True)
            print(f"  等待超时时间: {LOGIN_TIMEOUT} 秒", flush=True)
            print("  登录完成后脚本将自动继续...", flush=True)

            start = time.time()
            logged_in = False
            while time.time() - start < LOGIN_TIMEOUT:
                if not is_on_login_page(page):
                    logged_in = True
                    break
                elapsed = int(time.time() - start)
                if elapsed % 30 == 0 and elapsed > 0:
                    print(f"  已等待 {elapsed} 秒...", flush=True)
                page.wait_for_timeout(1000)

            if logged_in:
                print("[OK] 检测到已登录，继续采集...", flush=True)
                page.wait_for_timeout(5000)
            else:
                print("[!] 等待登录超时，将尝试直接采集...", flush=True)
        else:
            print("[OK] 已登录状态，直接开始采集...", flush=True)

        # 等待表格加载
        print("等待数据表格加载...", flush=True)
        if not wait_for_table(page):
            print("[!] 首次未检测到表格，等待5秒后重试...", flush=True)
            page.wait_for_timeout(5000)
            wait_for_table(page)

        # 开始采集
        all_rows = []
        page_num = 1

        while True:
            print(f"\n--- 正在抓取第 {page_num} 页 ---", flush=True)
            rows = scrape_page(page)

            if page_num == 1:
                all_rows.extend(rows)
            else:
                if rows and len(rows) > 1:
                    all_rows.extend(rows[1:])
                elif rows:
                    all_rows.extend(rows)

            print(f"  当前累计 {len(all_rows)} 行数据", flush=True)

            if not rows:
                print("  当前页无数据，采集结束", flush=True)
                break

            # 翻页
            try:
                next_selectors = [
                    ".ant-pagination-next:not(.ant-pagination-disabled)",
                    ".el-pagination button.btn-next:not([disabled])",
                    "li.next:not(.disabled)",
                    "button:has-text('下一页'):not([disabled])",
                    ".pagination .next:not(.disabled)",
                    "[class*=pagination] .next:not(.disabled)",
                ]
                next_btn = None
                for sel in next_selectors:
                    next_btn = page.query_selector(sel)
                    if next_btn and next_btn.is_visible():
                        break

                if next_btn:
                    next_btn.click()
                    page.wait_for_timeout(3000)
                    page_num += 1
                else:
                    print("  没有找到可用的翻页按钮，采集结束", flush=True)
                    break
            except Exception as e:
                print(f"  翻页结束: {e}", flush=True)
                break

        # 截图保存
        try:
            screenshot_path = os.path.join(DESKTOP, "page_screenshot.png")
            page.screenshot(path=screenshot_path, full_page=True)
            print(f"\n页面截图已保存到: {screenshot_path}", flush=True)
        except Exception:
            pass

        # 先保存已采集的数据
        if all_rows:
            save_to_excel(all_rows, OUTPUT_FILE)
        else:
            print("\n[!] 未采集到表格数据", flush=True)

        # 保持浏览器打开，等待用户操作
        print("\n" + "=" * 60, flush=True)
        print("  浏览器保持打开，你可以继续在页面上操作。", flush=True)
        print("  完成后关闭浏览器窗口即可自动退出。", flush=True)
        print("=" * 60, flush=True)

        try:
            while True:
                page.wait_for_timeout(1000)
                try:
                    page.title()
                except Exception:
                    print("浏览器已关闭，退出。", flush=True)
                    break
        except KeyboardInterrupt:
            print("手动退出。", flush=True)


if __name__ == "__main__":
    main()
