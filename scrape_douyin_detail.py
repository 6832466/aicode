"""抓取短剧版权平台 - 原生素材详情数据"""
import os, sys, time, re

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from playwright.sync_api import sync_playwright
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

LIST_URL = "https://www.shortdramas.com/page/copyright/book-manage?tab=motion_comic"
DESKTOP = os.path.join(os.path.expanduser("~"), "Desktop")
USER_DATA_DIR = os.path.join(os.environ.get("TEMP", "C:/Temp"), "playwright_chrome_profile")

# 目标作品名和输出文件名（从命令行参数读取，或使用默认）
TARGET_NAME = sys.argv[1] if len(sys.argv) > 1 else "我开废品收购不丢人"
OUTPUT_FILE = os.path.join(DESKTOP, f"{TARGET_NAME}.xlsx")


def click_text(page, text, timeout=10000):
    """点击包含指定文本的元素"""
    try:
        elem = page.wait_for_selector(f"text='{text}'", timeout=timeout, state="visible")
        elem.click()
        print(f"  [OK] 点击了 '{text}'", flush=True)
        return True
    except Exception as e:
        print(f"  [X] 未找到 '{text}': {e}", flush=True)
        return False


def select_only_category(page, target_cat):
    """点'全部'让所有选中，再取消不需要的分类，只保留目标"""
    # Step 1: 点击"全部"——所有标签变为选中
    try:
        all_span = page.query_selector(
            "[class*=meterialTypeTab] span:has-text('全部')"
        )
        if all_span:
            all_span.click()
            page.wait_for_timeout(600)
            print(f"    点击了全部", flush=True)
    except Exception as e:
        print(f"    [!] 点击全部失败: {e}", flush=True)

    page.wait_for_timeout(800)

    # Step 2: 取消不需要的分类（点击已选中的来取消）
    all_cats = ["高光-剧情回顾", "高光-剧情解析", "高光-预告片", "花絮", "高光-其他"]
    for cat in all_cats:
        if cat == target_cat:
            continue  # 保留目标
        try:
            span = page.query_selector(
                f"span.semi-checkbox-checked.semi-checkbox-cardType:has-text('{cat}')"
            )
            if span:
                span.click()
                page.wait_for_timeout(400)
                print(f"    取消: {cat}", flush=True)
        except Exception as e:
            pass  # 可能已经被取消或其他原因

    page.wait_for_timeout(2000)
    return True


def handle_popup(page):
    """关闭弹窗"""
    for text in ["我知道了", "知道了", "确定", "确认"]:
        try:
            btn = page.query_selector(f"button:has-text('{text}')")
            if btn and btn.is_visible():
                btn.click()
                print(f"  [OK] 关闭弹窗 '{text}'", flush=True)
                page.wait_for_timeout(1500)
                return True
        except Exception:
            pass
    return False


def scrape_video_cards(page):
    """抓取页面中所有视频卡片数据"""
    page.wait_for_timeout(3000)
    cards_data = []

    # 多种方式查找卡片容器
    cards = page.query_selector_all("[class*=positive-card-s]")
    if not cards:
        cards = page.query_selector_all("[class*=sortable-item-container]")
    if not cards:
        cards = page.query_selector_all("[class*=positiveVideos] > div")
    if not cards:
        cards = page.query_selector_all("[class*=content-positive-list] > div > div")

    # DEBUG: 如果还是没找到，打印页面中的关键 class
    if not cards:
        print(f"    [DEBUG] 未找到卡片，搜索页面结构...", flush=True)
        all_divs = page.query_selector_all("div")
        found_classes = set()
        for div in all_divs:
            try:
                cls = div.get_attribute("class") or ""
                for part in cls.split():
                    if "positive" in part.lower() or "card" in part.lower() or "sortable" in part.lower():
                        found_classes.add(part)
            except Exception:
                pass
        if found_classes:
            print(f"    找到的相关 class: {list(found_classes)[:20]}", flush=True)
        else:
            # 最后尝试：打印可见文本
            body_text = page.inner_text("body")
            print(f"    页面可见文本 (前500字): {body_text[:500]}", flush=True)

    print(f"    找到 {len(cards)} 个视频卡片", flush=True)

    for card in cards:
        try:
            data = {}

            # 标题
            title_el = card.query_selector("[class*=positive-card-content-title]:not([class*=block])")
            if title_el:
                data["标题"] = title_el.inner_text().strip()

            # 类型标签
            type_el = card.query_selector("[class*=positive-card-cover-material-type]")
            if type_el:
                data["类型"] = type_el.inner_text().strip()

            # 时长
            dur_el = card.query_selector("[class*=positive-card-cover-duration]")
            if dur_el:
                data["时长"] = dur_el.inner_text().strip()

            # 状态
            status_el = card.query_selector("[class*=positive-card-content-status]")
            if status_el:
                data["发布状态"] = status_el.inner_text().strip()

            # 时间
            time_el = card.query_selector("[class*=positive-card-content-time]")
            if time_el:
                data["时间"] = time_el.inner_text().strip()

            # 数据项（浏览/评论/点赞）
            data_items = card.query_selector_all("[class*=positive-card-content-data-item]")
            labels = ["浏览数", "评论数", "点赞数"]
            for i, item in enumerate(data_items):
                if i < len(labels):
                    data[labels[i]] = item.inner_text().strip()

            # 序号
            order_el = card.query_selector("[class*=positive-card-order]")
            if order_el:
                data["序号"] = order_el.inner_text().strip()

            # 用类型+序号作为标题
            if data.get("类型") and data.get("序号"):
                data["标题"] = f"{data['类型']} #{data['序号']}"

            # 过滤空卡片/占位符（无标题、无类型、含"暂无"）
            has_real_data = (
                data.get("类型") or data.get("时长") or data.get("浏览数")
            )
            card_text = card.inner_text()
            is_placeholder = "暂无" in card_text

            if has_real_data and not is_placeholder:
                cards_data.append(data)
        except Exception as e:
            print(f"    卡片解析异常: {e}", flush=True)

    return cards_data


def save_to_excel(all_data, filepath):
    """保存到 Excel，每个分类一个 Sheet"""
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    cell_align = Alignment(vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 标准列顺序
    columns = ["序号", "标题", "类型", "时长", "发布状态", "时间", "浏览数", "评论数", "点赞数"]

    total = 0
    for cat_name, cards in all_data.items():
        if not cards:
            continue

        # Sheet名最多31字符
        sheet_name = cat_name.replace("高光-", "").replace("/", "-")[:31]
        ws = wb.create_sheet(title=sheet_name)

        # 表头
        for ci, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=ci, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # 数据行
        for ri, card in enumerate(cards, 2):
            for ci, col_name in enumerate(columns, 1):
                val = card.get(col_name, "")
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = cell_align
                cell.border = thin_border
                cell.font = Font(name="微软雅黑", size=10)
            total += 1

        # 列宽
        col_widths = [8, 40, 18, 10, 12, 22, 12, 12, 12]
        for ci, w in enumerate(col_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    # 汇总 Sheet
    if len(all_data) > 1:
        summary_ws = wb.create_sheet(title="汇总", index=0)
        for ci, col_name in enumerate(columns, 1):
            cell = summary_ws.cell(row=1, column=ci, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        row_idx = 2
        for cat_name, cards in all_data.items():
            for card in cards:
                for ci, col_name in enumerate(columns, 1):
                    val = card.get(col_name, "")
                    cell = summary_ws.cell(row=row_idx, column=ci, value=val)
                    cell.alignment = cell_align
                    cell.border = thin_border
                    cell.font = Font(name="微软雅黑", size=10)
                row_idx += 1

        col_widths = [8, 40, 18, 10, 12, 22, 12, 12, 12]
        for ci, w in enumerate(col_widths, 1):
            summary_ws.column_dimensions[summary_ws.cell(row=1, column=ci).column_letter].width = w

    wb.save(filepath)
    print(f"\n[数据] 已保存到: {filepath}", flush=True)
    print(f"   共 {total} 条视频记录", flush=True)


def main():
    print("=" * 60, flush=True)
    print("  短剧版权平台 - 原生素材详情采集", flush=True)
    print("=" * 60, flush=True)
    os.makedirs(USER_DATA_DIR, exist_ok=True)

    all_data = {}
    categories = ["高光-剧情回顾", "高光-剧情解析", "高光-预告片", "花絮"]

    with sync_playwright() as p:
        context = p.chromium.launch_persistent_context(
            user_data_dir=USER_DATA_DIR,
            headless=False,
            args=["--start-maximized"],
            no_viewport=True,
        )
        page = context.new_page()
        page.set_default_timeout(15000)

        # ===== Step 1: 打开列表页 =====
        print(f"\n[Step 1] 打开列表页...", flush=True)
        try:
            page.goto(LIST_URL, wait_until="domcontentloaded", timeout=60000)
        except Exception:
            pass
        page.wait_for_timeout(5000)
        print(f"  URL: {page.url[:100]}", flush=True)

        # 检测是否在登录页
        if "login" in page.url.lower() or "signin" in page.url.lower():
            print("\n[!] 检测到登录页面，请在浏览器中完成登录...", flush=True)
            print("  登录完成后脚本将自动继续...", flush=True)
            start = time.time()
            while time.time() - start < 600:
                if "login" not in page.url.lower():
                    print("[OK] 登录完成，继续...", flush=True)
                    page.wait_for_timeout(3000)
                    break
                elapsed = int(time.time() - start)
                if elapsed % 30 == 0 and elapsed > 0:
                    print(f"  已等待 {elapsed} 秒...", flush=True)
                page.wait_for_timeout(1000)
            else:
                print("[!] 等待登录超时", flush=True)

        # ===== Step 2: 查找目标并进入详情页 =====
        print(f"\n[Step 2] 查找目标作品 '{TARGET_NAME}'...", flush=True)

        # 等待数据表格渲染（多种方式 + 额外延迟）
        page.wait_for_timeout(8000)
        row_sel = None
        for sel in [
            "table tbody tr",
            ".ant-table-row",
            ".arco-table-tr",
            "tr[class*='row']",
        ]:
            try:
                page.wait_for_selector(sel, timeout=5000)
                row_sel = sel
                break
            except Exception:
                continue

        page.wait_for_timeout(2000)

        found = False
        trs = page.query_selector_all(row_sel or "table tbody tr")
        print(f"  选择器: {row_sel}, 找到 {len(trs)} 行", flush=True)

        for tr in trs:
            try:
                row_text = tr.inner_text()
                if TARGET_NAME in row_text:
                    print(f"  找到目标行: {row_text[:120]}", flush=True)
                    links = tr.query_selector_all("a")
                    for link in links:
                        href = link.get_attribute("href") or ""
                        if "short-play-detail" in href:
                            detail_url = href if href.startswith("http") else f"https://www.shortdramas.com{href}"
                            print(f"  [OK] 跳转到: {detail_url[:100]}", flush=True)
                            page.goto(detail_url, wait_until="domcontentloaded", timeout=60000)
                            found = True
                            break
                    if found:
                        break
                    detail_btn = tr.query_selector("text='查看详情'")
                    if detail_btn:
                        detail_btn.click()
                        page.wait_for_timeout(5000)
                        found = True
                        break
            except Exception:
                continue

        if not found:
            # 再次尝试：直接遍历页面所有链接
            print("  遍历链接中...", flush=True)
            all_links = page.query_selector_all("a")
            for link in all_links:
                try:
                    href = link.get_attribute("href") or ""
                    link_text = link.inner_text().strip()
                    if TARGET_NAME in link_text or (
                        "short-play-detail" in href and TARGET_NAME in page.inner_text("body")
                    ):
                        print(f"  匹配链接: text='{link_text[:60]}' href='{href[:80]}'", flush=True)
                        if "short-play-detail" in href:
                            detail_url = href if href.startswith("http") else f"https://www.shortdramas.com{href}"
                            page.goto(detail_url, wait_until="domcontentloaded", timeout=60000)
                            found = True
                            break
                except Exception:
                    continue

        if not found:
            print(f"  [!] 未找到目标作品 '{TARGET_NAME}'，打印页面搜索信息:", flush=True)
            body_text = page.inner_text("body")
            if TARGET_NAME in body_text:
                idx = body_text.index(TARGET_NAME)
                print(f"    在页面中找到 '{TARGET_NAME}' (位置 {idx})", flush=True)
            else:
                print(f"    页面中未找到 '{TARGET_NAME}'", flush=True)
            print(f"    页面文本 (前300字): {body_text[:300]}", flush=True)
            print("[!] 请手动找到目标并点击查看详情，等待5分钟...", flush=True)
            page.wait_for_timeout(300000)

        page.wait_for_timeout(5000)
        print(f"  当前 URL: {page.url[:120]}", flush=True)

        # ===== Step 3: 点击"抖音原生经营" =====
        print(f"\n[Step 3] 点击'抖音原生经营'...", flush=True)
        if not click_text(page, "抖音原生经营"):
            print("  请手动点击 抖音原生经营...", flush=True)
            page.wait_for_timeout(60000)
        page.wait_for_timeout(4000)

        # ===== Step 4: 点击"原生素材" tab =====
        print(f"\n[Step 4] 点击'原生素材' tab...", flush=True)
        tab_clicked = False
        # 用 page.click（自带等待和重试），与探路脚本一致
        for sel in [
            ".arco-tabs-header-title:has-text('原生素材')",
            "text='原生素材'",
            "span:has-text('原生素材')",
        ]:
            try:
                page.click(sel, timeout=10000)
                tab_clicked = True
                print(f"  [OK] 点击了原生素材 tab: {sel}", flush=True)
                break
            except Exception:
                continue

        if not tab_clicked:
            print("  请手动点击 原生素材 tab...", flush=True)
            page.wait_for_timeout(60000)

        page.wait_for_timeout(4000)

        # ===== Step 5: 处理弹窗 =====
        print(f"\n[Step 5] 检查弹窗...", flush=True)
        for _ in range(3):
            handle_popup(page)
            page.wait_for_timeout(1000)
        page.wait_for_timeout(2000)
        handle_popup(page)

        # ===== Step 6: 截图当前状态 =====
        try:
            page.screenshot(path=os.path.join(DESKTOP, "detail_screenshot.png"), full_page=True)
            print("  截图已保存", flush=True)
        except Exception:
            pass

        # ===== Step 7: 逐个分类精确抓取 =====
        print(f"\n[Step 7] 逐分类抓取视频数据...", flush=True)

        for cat in categories:
            print(f"\n--- {cat} ---", flush=True)

            # 点全部 → 取消其他 → 只保留目标
            select_only_category(page, cat)
            page.wait_for_timeout(1000)

            # 处理弹窗
            handle_popup(page)
            page.wait_for_timeout(1000)

            # 抓取卡片
            cards = scrape_video_cards(page)
            all_data[cat] = cards
            print(f"  {cat}: {len(cards)} 条视频", flush=True)

        # ===== 导出 Excel =====
        if any(v for v in all_data.values()):
            save_to_excel(all_data, OUTPUT_FILE)
        else:
            print("\n[!] 未采集到数据", flush=True)

        # ===== 保持浏览器打开 =====
        print("\n" + "=" * 60, flush=True)
        print("  浏览器保持打开，你可以继续操作。", flush=True)
        print("  完成后关闭浏览器窗口即可退出。", flush=True)
        print("=" * 60, flush=True)

        try:
            while True:
                page.wait_for_timeout(1000)
                try:
                    page.title()
                except Exception:
                    print("浏览器已关闭，退出。", flush=True)
                    break
        except KeyboardInterrupt:
            print("手动退出。", flush=True)


if __name__ == "__main__":
    main()
